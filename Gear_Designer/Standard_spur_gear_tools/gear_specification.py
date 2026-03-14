import numpy as np
import openpyxl
import os

def pitch_diameter (M,T):#節徑(模數,齒數)
    Dm = T*M
    return(Dm)

def pitch_radius (M,T):#節圓半徑(模數,齒數)
    Rm = pitch_diameter (M,T)/2
    return(Rm)

def pitch_circumference (M,T):#節圓周長(模數,齒數)
    Cm = pitch_diameter (M,T)*np.pi
    return(Cm)

def tip_diameter (M,T):#齒頂圓半直徑(模數,齒數)
    Dt = (T+2)*M
    return(Dt)

def tip_radius (M,T):#齒頂圓半徑(模數,齒數)
    Rt = tip_diameter (M,T)/2
    return(Rt)

def root_diameter (M,T):#齒根圓直徑(模數,齒數)
    Dr = M*(T-2.5)
    return(Dr)

def root_radius (M,T):#齒根圓半徑(模數,齒數)
    Rr = root_diameter(M,T)/2
    return(Rr)

def base_diameter (M,T):#基圓直徑(模數,齒數)
    theta = 20#角度
    Db =  M*T*(np.cos(np.pi/(180/theta)))
    return(Db)

def base_radius (M,T):#基圓半徑(模數,齒數)
    Db = base_diameter (M,T)
    Rb = Db/2
    return(Rb)

def Addendum (M):#齒頂(模數)
    add = M
    return(add)

def Dedendum (M):#齒根(模數)
    Ded = 1.25*M
    return(Ded)

def tooth_height (M):#齒高(模數)
    h = 2.25*M
    return(h)

def working_depth (M):#工作深度(模數)
    Wd = M*2
    return(Wd)

def clearance (M):#齒輪間隙(模數)不是齒隙
    c = M*0.25
    return(c)

def circular_pitch (M):#周節(模數)
    cp = M*np.pi
    return(cp)

def tooth_thickness(M):#齒厚(模數)
    tt = circular_pitch (M)/2
    return(tt)

def tooth_space(M):#齒間(模數)
    ts = circular_pitch (M)/2
    return(ts)
    
def fillet(M):#齒根圓角半徑(模數)
    F = 0.236*M
    return(F)

def Diametral_pitch(M):#徑節(模數)
    Dp = 2.54/M
    return(Dp)

def pitch_angle (T):#周節角度(齒數)
    pa = 2*np.pi/T
    return(pa)

def halfpitch_angle (T):#半周節角度(齒數)
    pa = pitch_angle(T)
    hpa = pa/2
    return(hpa)

def gearBook(M, T, N):
    try:
        WB = openpyxl.load_workbook('gearbook.xlsx', data_only=True)
    except FileNotFoundError:
        WB = openpyxl.Workbook()  # 建立新的 Excel
        
    sheet = WB.active  # 使用預設的工作表
#===================================================(固定所有標題)
    title = [
             ['齒輪編號:'],['模數:'],['齒數:'],
             ['節圓直徑:'],['節圓半徑:'],['節圓周長:'],['齒頂圓直徑:'],['齒頂圓半徑:'],
             ['齒根圓直徑:'],['齒根圓半徑:'],['基圓直徑:'],['基圓半徑:'],['齒頂:'],['齒根:'],['齒高:'],['工作深度:'],
             ['間隙:'],['周節:'],['齒厚:'],['齒間:'],['齒根圓角半徑:'],['徑節:'],['壓力角:'],['每齒角度:'],['齒厚角度:']
             ]
    
    for y in range(len(title)):
        for x in range(len(title[y])):
            row = 2 + y      # 寫入資料的範圍從row(直向)開始
            col = 1 + x      # 寫入資料的範圍從column(橫向) 開始
            sheet.cell(row, col).value = title[y][x]
            
#===================================================(所有資料)
    data = [
            [N],[M],[T],
            [pitch_diameter(M,T)],[pitch_radius(M,T)],[pitch_circumference(M,T)],
            [tip_diameter(M,T)],[tip_radius(M,T)],[root_diameter(M,T)],
            [root_radius(M,T)],[base_diameter(M,T)],[base_radius(M,T)],[Addendum(M)],
            [Dedendum(M)],[tooth_height(M)],[working_depth(M)],
            [clearance(M)],[circular_pitch(M)],[tooth_thickness(M)],
            [tooth_space(M)],[fillet(M)],[Diametral_pitch(M)],[20],[pitch_angle(T)],[halfpitch_angle(T)]
            ]

    for y in range(len(data)):
        for x in range(len(data[y])):
            row = 2 + y      # 寫入資料的範圍從row(直向)開始
            col = N+1 + x      # 寫入資料的範圍從column(橫向) 開始
            sheet.cell(row, col).value = data[y][x]
        
    WB.save('gearbook.xlsx')#存檔

    
'''
M = 1
T = 10
print(gear_data(M,T,D))
gearBook(M,T,1)

'''

