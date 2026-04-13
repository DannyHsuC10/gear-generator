from pathlib import Path
import openpyxl
import os
import gear_parameters as gp


M = 1#模數(單位:mm)>0
T = 20#齒數>2
a = 1#齒頂倍率>0
d = 1.25#齒根倍率>0
s = 1#齒間倍率<2,>0
ap = 20#壓力角(單位:度)
f = 0.236#齒根圓角半徑倍率>0


# 使用相對路徑載入Excel檔案

PATH = Path(__file__).resolve().parent

def sorter (data,sheet,n):#資料排序(資料串列,清單編號,編號)
    for y in range(len(data)):
        for x in range(len(data[y])):
            row = 1 + y      # 寫入資料的範圍從row(直向)開始
            col = n + x      # 寫入資料的範圍從column(橫向) 開始
            sheet.cell(row, col).value = data[y][x]
    
def table (M,T,a,d,s,ap,f,N):#齒輪參數表格(模數,齒數,齒頂倍率,齒根倍率,齒間倍率,壓力角,齒根圓角倍率,小數點位數ㄝ
    print(os.getcwd())#印出位址
    print('gear.xlsx')#印出檔名
    print('Sheet')#印出表格名稱
    WB = openpyxl.load_workbook(PATH / 'gear.xlsx', data_only=True)
    sheet = WB['Sheet1']
#===================================================(所有資料)
    data_Basic = [
        ['基本參數',''],
        ['齒數:',f"{T:.{N}f}"],
        ['模數:',f"{M:.{N}f}"],
        ['導出參數',''],
        ['節圓直徑:',f"{gp.pitch_diameter(M,T):.{N}f}"],
        ['節圓半徑:',f"{gp.pitch_radius(M,T):.{N}f}"],
        ['節圓周長:',f"{gp.pitch_circumference(M,T):.{N}f}"],
        ['周節:',f"{gp.pitch(M):.{N}f}"],
        ['徑節:',f"{gp.Diametral_pitch(M):.{N}f}"],
        ['周節徑度:',f"{gp.pitch_radian (T):.{N}f}"]]
    
    data_Addendum = [
        ['頂部參數',''],
        ['頂部倍率:',f"{a:.{N}f}"],
        ['',''],
        ['導出參數',''],
        ['齒頂:',f"{gp.Addendum (M,a):.{N}f}"],
        ['齒頂圓直徑:',f"{gp.tip_diameter (M,T,a):.{N}f}"],
        ['齒頂圓半徑:',f"{gp.tip_radius (M,T,a):.{N}f}"],
        ['齒頂圓周長:',f"{gp.tip_circumference(M,T,a):.{N}f}"],
        ['工作深度:',f"{gp.working_depth (M,a):.{N}f}"]]
    
    data_Dedendum = [
        ['根部參數',''],
        ['根部倍率:',f"{d:.{N}f}"],
        ['',''],
        ['導出參數',''],
        ['齒根:',f"{gp.Dedendum (M,d):.{N}f}"],
        ['齒根圓直徑:',f"{gp.root_diameter (M,T,d):.{N}f}"],
        ['齒根圓半徑:',f"{gp.root_radius (M,T,d):.{N}f}"],
        ['齒根圓周長:',f"{gp.root_circumference (M,T,d):.{N}f}"],
        ['齒高:',f"{gp.tooth_height (M,a,d):.{N}f}"],
        ['齒隙:',f"{gp.clearance (M,a,d):.{N}f}"]]
    
    data_thickness = [
        ['厚度參數',''],
        ['齒間倍率:',f"{s:.{N}f}"],
        ['',''],
        ['導出參數',''],
        ['齒厚:',f"{gp.tooth_thickness(M,s):.{N}f}"],
        ['齒間:',f"{gp.tooth_space(M,s):.{N}f}"],
        ['齒厚徑度:',f"{gp.thickness_radian(M,T,s):.{N}f}"],
        ['齒間徑度:',f"{gp.space_radian(M,T,s):.{N}f}"]]
    
    data_pressure = [
        ['軸向分力參數',''],
        ['壓力角:',f"{ap:.{N}f}"],
        ['',''],
        ['導出參數',''],
        ['壓力角徑度:',f"{gp.pressure_radian (ap):.{N}f}"],
        ['基圓直徑:',f"{gp.base_diameter (M,T,ap):.{N}f}"],
        ['基圓半徑:',f"{gp.base_radius (M,T,ap):.{N}f}"],
        ['基圓周長:',f"{gp.base_circumference(M,T,ap):.{N}f}"]]
    
    data_fillet = [
        ['圓角參數',''],
        ['齒根圓角倍率:',f"{f:.{N}f}"],
        ['',''],
        ['導出參數',''],
        ['齒根圓角半徑:',f"{gp.fillet(M,f):.{N}f}"]]

    n = 1
    sorter (data_Basic,sheet,n)
    n = n+2
    sorter (data_Addendum,sheet,n)
    n = n+2
    sorter (data_Dedendum,sheet,n)
    n = n+2
    sorter (data_thickness,sheet,n)
    n = n+2
    sorter (data_pressure,sheet,n)
    n = n+2
    sorter (data_fillet,sheet,n)
    
    WB.save(PATH / 'gear.xlsx')#存檔

'''
資料編號1~6
1.基本資料
2.齒頂資料
3.齒根資料
4.齒間資料
5.壓力資料
6.圓角資料
基本參數項目1~3
導出參數項目4~10
'''

def get_data(n,s,x):#取得資料(資料編號,起始參數,最終參數)
    WB = openpyxl.load_workbook(PATH / 'gear.xlsx', data_only=True)
    sheet = WB['Sheet1']
    Data = []
    while s <= x:
        for y in range(2):
            text = sheet.cell(s, y+2*(n-1)+1).value
            if text == None:
                pass
            else:
                Data.append(text)
        s = s+1
    return(Data)
    

def get_value(x,y):
    WB = openpyxl.load_workbook(PATH / 'gear.xlsx', data_only=True)
    sheet = WB['Sheet1']
    cell_value = sheet.cell(x, y).value
    if cell_value is not None:
        value = float(cell_value)
    else:
        value = 0.0
    return(value)

def writer(x,y,n):
    WB = openpyxl.load_workbook('gear.xlsx', data_only=True)
    sheet = WB['Sheet1']
    sheet.cell(x, y).value = n
    


table (M,T,a,d,s,ap,f,3)
get_data(1,1,3)
get_data(1,4,10)
